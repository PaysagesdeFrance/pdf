#!/usr/bin/env python3
"""Régénère plu et rlp depuis l'export national BANATIC (intercommunalités)."""

import io
import re
import sys
import unicodedata

import requests
from openpyxl import load_workbook

URL = ("https://www.banatic.interieur.gouv.fr/consultation/api/"
       "export/pregenere/telecharger/France")

# Colonnes repérées par libellé normalisé — JAMAIS par lettre (DU, CH se décalent
# à chaque compétence ajoutée/retirée par la DGCL).
LABEL_SIREN = "nsiren"                       # « N° SIREN »
LABEL_PLU   = "planlocaldurbanisme"          # début du libellé Art. L. 153-1
LABEL_RLP   = "reglementlocaldepublicite"    # « Règlement local de publicité »

MIN_LIGNES = 1200   # ~1250 EPCI à fiscalité propre ; en dessous → export suspect


def norm(v):
    """Minuscules, sans accents, alphanumérique seul — même esprit que le norm() JS."""
    s = unicodedata.normalize("NFD", str(v or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s.lower())


def trouver_feuille_et_entete(wb):
    for ws in wb.worksheets:
        for i, row in enumerate(
                ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            cells = [norm(c) for c in row]
            if LABEL_SIREN in cells:
                return ws, i, cells
    sys.exit("Colonne « N° SIREN » introuvable : format BANATIC modifié ?")


def index_colonne(entete, prefixe):
    hits = [i for i, h in enumerate(entete) if h.startswith(prefixe)]
    if len(hits) != 1:
        sys.exit(f"Colonne « {prefixe} » : {len(hits)} correspondance(s) au lieu de 1. "
                 "Format BANATIC probablement modifié — fichiers NON écrits.")
    return hits[0]


def main():
    r = requests.get(URL, timeout=180,
                     headers={"User-Agent": "Mozilla/5.0 (MAJ PLU/RLP)"})
    r.raise_for_status()

    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws, ligne_entete, entete = trouver_feuille_et_entete(wb)

    i_siren = entete.index(LABEL_SIREN)
    i_plu   = index_colonne(entete, LABEL_PLU)
    i_rlp   = index_colonne(entete, LABEL_RLP)

    plu, rlp, rejets = [], [], 0
    for row in ws.iter_rows(min_row=ligne_entete + 1, values_only=True):
        siren = re.sub(r"\D", "", str(row[i_siren] or ""))
        if not re.fullmatch(r"\d{9}", siren):
            continue                                   # ligne vide, total, etc.
        for i_col, dest in ((i_plu, plu), (i_rlp, rlp)):
            v = norm(row[i_col])
            if v == "oui":
                dest.append(f"{siren},1")
            elif v == "non":
                dest.append(f"{siren},0")
            else:
                rejets += 1                            # valeur inattendue → omise

    if min(len(plu), len(rlp)) < MIN_LIGNES:
        sys.exit(f"Trop peu de lignes (plu={len(plu)}, rlp={len(rlp)}) : "
                 "format probablement modifié — fichiers NON écrits.")
    if rejets:
        print(f"Avertissement : {rejets} valeur(s) ni OUI ni NON ignorée(s).")

    for nom, lignes in (("plu", plu), ("rlp", rlp)):
        with open(nom, "w", encoding="utf-8", newline="\n") as f:
            f.write("\n".join(lignes) + "\n")
        print(f"{nom} : {len(lignes)} lignes.")


if __name__ == "__main__":
    main()
